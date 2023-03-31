import numpy as np
from openpyxl import load_workbook, Workbook

ws = load_workbook("log data.xlsx")
ws = ws['Sheet1']
max_row = ws.max_row - 2

data = []
for i in range(1, max_row + 1) :
    data.append([ ws[f"A{i}"].value , ws[f"B{i}"].value ])

data = np.array(data)

x = data[:, 0]
y = data[:, 1]

dy = y[1:] - y[:-1]
dx = x[1:] - x[:-1]
deriv = dy/dx

ddy = dx.copy()
ddy[0] = dy[0]
ddy[1:] = dy[1:] - dy[:-1]
second_deriv = y.copy()
second_deriv[1:] = ddy/dx
second_deriv[0] = y[0]

# write data in an elsx file :
wb_save = Workbook()               
ws_save = wb_save.active

i = 0
for x_i in x :
    ws_save[f"A{i+1}"] = x_i
    ws_save[f"B{i+1}"] = second_deriv[i]
    i += 1

wb_save.save("second_deriv.xlsx")
